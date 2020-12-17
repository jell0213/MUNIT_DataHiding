# -*- coding: utf-8 -*-
#######################################################
'''
input
    資料夾名稱
    圖片數量
    MOD值
    嵌密率
處理內容
    輸入一張圖片的資料，包含：
        1.資料夾名稱
        2.檔案名稱(圖片)，單純用來記錄在xlsx檔案中
        3.輸出路徑-xlsx
        4.輸出路徑-lc-txt
        5.輸出路徑-lc-png
        6.嵌密mod值
        7.輸出路徑-lc-txt-壓縮
        8.輸出路徑-lc-png-壓縮
        9.嵌密率
output
    產生輸入圖片的xlsx檔(依序將所有圖片的資料寫入xlsx檔中)      
        包含執行時間
注意事項
    目前路徑固定為embed_mod3資料夾的相對路徑(本程式與此資料夾要在相同路徑)
'''
#######################################################
from skimage import io
from openpyxl import Workbook
import openpyxl
import os
import math
import time
def cal_capacity(folder_name,
                 num_image,
                 num_mod,
                 embed_ratio):
    wb = Workbook()
    ws = wb.active
    ws.append(["embed_mod3","mod="+str(num_mod),str(embed_ratio)+"%","256*256"])
    ws.append(["","","文字檔","","","","","","","圖片檔"])
    ws.append(["檔名","嵌密量","大小(bit)","壓縮大小","壓縮率","嵌密壓縮率","淨藏量","rare bpp","bpp","大小(bit)","壓縮大小","壓縮率","嵌密壓縮率","淨藏量","pure bpp","bpp"])
    wb.save("./embed_mod3/embed_mod3_capacity.xlsx")
    wb = openpyxl.load_workbook("./embed_mod3/embed_mod3_capacity.xlsx")
    ws = wb['Sheet']
    a=[]#儲存各項平均值
    for i in range(20):
        a.append(0) 
    for i in range(num_image):
        f_lc= open("./embed_mod3/output{:08d}".format(i)+"/output{:08d}_location map.txt".format(i),'r')                                      #打開location map.txt來計算capacity
        image_lc = io.imread("./embed_mod3/output{:08d}".format(i)+"/output{:08d}_lc.png".format(i))
        count = 0    
        for row in range(image_lc.shape[0]):                        #計算capacity，非白區域可嵌密，三個channel，對mod以2為底取log(單位：bit)
            for col in range(image_lc.shape[1]):
                bit=f_lc.read(1)
                if bit== "0" or bit== "2" :
                    count+=1
        count*=3*math.log(num_mod,2)*embed_ratio/100                                #capacity        
        size_lc=os.path.getsize("./embed_mod3/output{:08d}".format(i)+"/output{:08d}_location map.txt".format(i))*8#txt大小
        size_image_lc=os.path.getsize("./embed_mod3/output{:08d}".format(i)+"/output{:08d}_lc.png".format(i))*8#png大小
        size_lc_gz=os.path.getsize("./embed_mod3/output{:08d}".format(i)+"/output{:08d}_location map.tar.gz".format(i))*8#txt壓縮大小
        size_image_lc_gz=os.path.getsize("./embed_mod3/output{:08d}".format(i)+"/output{:08d}_lc.tar.gz".format(i))*8#png壓縮大小
        compress_lc=(size_lc_gz/size_lc)*100#lc壓縮率
        compress_image_lc=(size_image_lc_gz/size_image_lc)*100#lm壓縮率
        compress_code_lc=(size_lc_gz/count)*100#lc嵌密壓縮率(%)
        compress_code_image_lc=(size_image_lc_gz/count)*100#lm嵌密壓縮率(%)
        net_capacity_lc=count-size_lc_gz#lc.txt淨藏量
        net_capacity_image_lc=count-size_image_lc_gz#lc.png淨藏量
        net_embedding_ratio_lc=net_capacity_lc/(256*256)#lc.txt淨嵌入率(%)
        net_embedding_ratio_image_lc=net_capacity_image_lc/(256*256)#lc.png淨嵌入率(%)
        embedding_ratio_lc=count/(256*256)#嵌入率(%)(txt和png相同)
        
        ws.append(["output{:08d}".format(i),
                   float('%.2f'%round(count,2)),    #四捨五入到指定小數位
                   float('%.2f'%round(size_lc,2)),
                   float('%.2f'%round(size_lc_gz,2)),
                   str(float('%.1f'%round(compress_lc,1)))+'%',
                   str(float('%.1f'%round(compress_code_lc,1)))+'%',
                   float('%.2f'%round(net_capacity_lc,2)),
                   float('%.2f'%round(net_embedding_ratio_lc,2)),
                   float('%.2f'%round(embedding_ratio_lc,2)),
                   float('%.2f'%round(size_image_lc,2)),
                   float('%.2f'%round(size_image_lc_gz,2)),
                   str(float('%.1f'%round(compress_image_lc,1)))+'%',
                   str(float('%.1f'%round(compress_code_image_lc,1)))+'%',
                   float('%.2f'%round(net_capacity_image_lc,2)),
                   float('%.2f'%round(net_embedding_ratio_image_lc,2)),
                   float('%.2f'%round(embedding_ratio_lc,2))])
        a[0]+=count
        a[1]+=size_lc
        a[2]+=size_lc_gz
        a[3]+=compress_lc
        a[4]+=compress_code_lc
        a[5]+=net_capacity_lc
        a[6]+=net_embedding_ratio_lc
        a[7]+=embedding_ratio_lc
        a[8]+=size_image_lc
        a[9]+=size_image_lc_gz
        a[10]+=compress_image_lc
        a[11]+=compress_code_image_lc
        a[12]+=net_capacity_image_lc
        a[13]+=net_embedding_ratio_image_lc
        a[14]+=embedding_ratio_lc
        f_lc.close()
    
    for i in range(20):
        a[i]/=num_image
    ws.append(["檔名","嵌密量","大小(bit)","壓縮大小","壓縮率","嵌密壓縮率","淨藏量","pure bpp","bpp","大小(bit)","壓縮大小","壓縮率","嵌密壓縮率","淨藏量","pure bpp","bpp"])
    ws.append([
            "",
            float('%.2f'%round(a[0],2)),
            float('%.2f'%round(a[1],2)),
            float('%.2f'%round(a[2],2)),
            str(float('%.1f'%round(a[3],1)))+'%',
            str(float('%.1f'%round(a[4],1)))+'%',
            float('%.2f'%round(a[5],2)),
            float('%.2f'%round(a[6],2)),
            float('%.2f'%round(a[7],2)),
            float('%.2f'%round(a[8],2)),
            float('%.2f'%round(a[9],2)),
            str(float('%.1f'%round(a[10],1)))+'%',
            str(float('%.1f'%round(a[11],1)))+'%',
            float('%.2f'%round(a[12],2)),
            float('%.2f'%round(a[13],2)),
            float('%.2f'%round(a[14],2)),
            ])
    wb.save("./embed_mod3/embed_mod3_capacity.xlsx")#寫檔後存檔
embed_ratio=int(input("embedding ratio(%) = "))
tStart = time.time()#計時開始
cal_capacity("embed_mod3",5000,3,embed_ratio)
tEnd = time.time()#計時結束
wb = openpyxl.load_workbook("./embed_mod3/embed_mod3_capacity.xlsx")
ws = wb['Sheet']
ws.append(["total time",str(round(tEnd-tStart,2))+" s"])
wb.save("./embed_mod3/embed_mod3_capacity.xlsx")#寫檔後存檔
print(round(tEnd-tStart,2))
