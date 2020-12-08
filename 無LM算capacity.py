# -*- coding: utf-8 -*-
#######################################################
'''
input
    資料夾名稱
    圖片數量
    MOD值
    嵌密率
處理內容
    輸入一張圖片的資料，包含：
        1.資料夾名稱
        2.檔案名稱(圖片)，單純用來記錄在xlsx檔案中
        3.輸出路徑-xlsx
        4.嵌密mod值
        5.嵌密率
output
    產生輸入圖片的xlsx檔(依序將所有圖片的資料寫入xlsx檔中)      
        包含執行時間
注意事項
    目前路徑固定為embed_mod3資料夾的相對路徑(本程式與此資料夾要在相同路徑)
'''
#######################################################
from skimage import io
from openpyxl import Workbook
import openpyxl
import os
import math
import time
def cal_capacity(in_dir,
                 num_image,
                 num_mod,
                 embed_ratio):
    wb = Workbook()
    ws = wb.active
    ws.append(["無LM","mod="+str(num_mod),str(embed_ratio)+"%","256*256"])
    ws.append(["檔名","嵌密量","bpp"])
    wb.save(in_dir+"/embed_mod3_capacity.xlsx")
    wb = openpyxl.load_workbook(in_dir+"/embed_mod3_capacity.xlsx")
    ws = wb['Sheet']
    a=[]                                                                                                    #儲存各項平均值
    for i in range(2):
        a.append(0)
    for i in range(num_image):
        f_code= in_dir+"/output{:08d}".format(i)+"/output{:08d}_code.txt".format(i)    #打開location map.txt來計算capacity
        #count = 0    
        try:#計算code.txt的字數(計算嵌了幾個bit)
            with open(f_code, encoding='utf8') as file_obj:
            #由於書名不是英文，要加上 encoding='utf8'
                contents = file_obj.read()
            
        except FileNotFoundError:
            msg = 'Sorry, the file' + f_code + ' does not exist.'
            print(msg)
        else:
            words = contents.rstrip()
            num_words = len(words)   
            print (1)
        num_words*=3*math.log(num_mod,2)*embed_ratio/100                                                        #capacity 
        embedding_ratio_lc=num_words/(256*256)                                                                  #嵌入率(%)(txt和png相同)        
        ws.append(["output{:08d}".format(i),
                   float('%.2f'%round(num_words,2)),                                                            #四捨五入到指定小數位
                   float('%.2f'%round(embedding_ratio_lc,2))])
        a[0]+=num_words
        a[1]+=embedding_ratio_lc
        if i%250 == 0:
            print(i)
    for i in range(2):
        a[i]/=num_image
    ws.append(["檔名","嵌密量","bpp"])
    ws.append([
            "",
            float('%.2f'%round(a[0],2)),            
            float('%.2f'%round(a[1],2)),
            ])
    wb.save(in_dir+"/embed_mod3_capacity.xlsx")                                                        #寫檔後存檔
embed_ratio=int(input("embedding ratio(%) = "))
tStart = time.time()                                                                                        #計時開始
#in_dir="D:\\108resercher\\====######RESEARCH######====\\GAN-research\\12.8\\無LM嵌密結果\\100%MOD3"
in_dir="D:\\108resercher\\====######RESEARCH######====\\GAN-research\\12.8\\無LM嵌密結果\\50%MOD3"
cal_capacity(in_dir,5000,3,embed_ratio)                                                               #執行程式------------------------------------------------------------
tEnd = time.time()                                                                                          #計時結束
wb = openpyxl.load_workbook(in_dir+"/embed_mod3_capacity.xlsx")
ws = wb['Sheet']
ws.append(["total time",str(round(tEnd-tStart,2))+" s"])
wb.save(in_dir+"/embed_mod3_capacity.xlsx")                                                            #寫檔後存檔
print(round(tEnd-tStart,2))
