#!/usr/bin/env python
# -*- coding: utf-8 -*-
#-----------------------------
#log
#檢測5000張256*256影像------>8小時50分


#-----------------------------
import cv2
from os.path import join
import os
from openpyxl import Workbook, load_workbook
import numpy as np
from skimage import io,color,img_as_ubyte
def pic_compare(name1,name2):
    image = io.imread(name1)
    image2 = io.imread(name2)
    difference = cv2.subtract(image,image2)
    result = not np.any(difference)
    if result is True :
        return 1
    else:
        return 0
    
    
    
    
count = 0
total = 5000
path = 'C:/Users/li/Desktop/stego_small'
for i in range(total):
    if i % 30 == 0 :
        print(i)
    for j in range(total-i):
        count+=pic_compare(path+"/output{:08d}_small.png".format(i),path+"/output{:08d}_small.png".format(total-j-1))

path_excel = join(path,"相同分析.xlsx")
if not os.path.isfile(path_excel) :
    wb = Workbook()
    ws = wb.active
    wb.save(path_excel)
wb = load_workbook(path_excel)
ws = wb.active
ws.append([count-total])
wb.save(path_excel)
print(count-total)