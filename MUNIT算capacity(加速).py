# -*- coding: utf-8 -*-
#######################################################
'''
input
    資料夾名稱
    圖片數量
    MOD值
    嵌密率
處理內容
    輸入一張圖片的資料，包含：
        1.資料夾名稱
        2.檔案名稱(圖片)，單純用來記錄在xlsx檔案中
        3.輸出路徑-xlsx
        4.輸出路徑-lc-txt
        5.輸出路徑-lc-png
        6.嵌密mod值
        7.輸出路徑-lc-txt-壓縮
        8.輸出路徑-lc-png-壓縮
        9.嵌密率
output
    產生輸入圖片的xlsx檔(依序將所有圖片的資料寫入xlsx檔中)        
'''
#######################################################
from skimage import io
from openpyxl import Workbook
import openpyxl
import os
import math
import time
def cal_capacity(folder_name,
                 num_image,
                 num_mod,
                 embed_ratio):
    try:#若檔案存在，直接打開繼續寫xlsx檔
        wb = openpyxl.load_workbook("./embed_mod3/embed_mod3_capacity.xlsx")
        #wb = openpyxl.load_workbook(dir_out)
    except:#若檔案不存在，建立新xlsx檔
        wb = Workbook()
        ws = wb.active
        ws.append(["embed_mod3","mod="+str(num_mod)])
        ws.append(["","","文字檔","","","","圖片檔"])
        ws.append(["檔名","嵌密量","大小(bit)","壓縮大小","檔案壓縮","嵌密壓縮","大小(bit)","壓縮大小","檔案壓縮","嵌密壓縮"])
        wb.save("./embed_mod3/embed_mod3_capacity.xlsx")
        wb = openpyxl.load_workbook("./embed_mod3/embed_mod3_capacity.xlsx")
    ws = wb['Sheet']
    for i in range(num_image):
        f_lc= open("./embed_mod3/output{:08d}".format(i)+"/output{:08d}_location map.txt".format(i),'r')                                      #打開location map.txt來計算capacity
        image_lc = io.imread("./embed_mod3/output{:08d}".format(i)+"/output{:08d}_lc.png".format(i))
        count = 0    
        
        for row in range(image_lc.shape[0]):                        #計算capacity，非白區域可嵌密，三個channel，對mod以2為底取log(單位：bit)
            for col in range(image_lc.shape[1]):
                bit=f_lc.read(1)
                if bit== "0" or bit== "2" :
                    count+=1
        count*=3*math.log(num_mod,2)*embed_ratio/100                                #capacity        
        size_lc=os.path.getsize("./embed_mod3/output{:08d}".format(i)+"/output{:08d}_location map.txt".format(i))*8#lc(txt)大小
        size_image_lc=os.path.getsize("./embed_mod3/output{:08d}".format(i)+"/output{:08d}_lc.png".format(i))*8#lm(png)大小
        size_lc_gz=os.path.getsize("./embed_mod3/output{:08d}".format(i)+"/output{:08d}_location map.tar.gz".format(i))*8#lc壓縮大小
        size_image_lc_gz=os.path.getsize("./embed_mod3/output{:08d}".format(i)+"/output{:08d}_lc.tar.gz".format(i))*8#lm壓縮大小
        compress_lc=((size_lc-size_lc_gz)/size_lc)*100
        compress_image_lc=((size_image_lc-size_image_lc_gz)/size_image_lc)*100
        compress_code_lc=((count-size_lc_gz)/count)*100#lc嵌密壓縮率(%)
        compress_code_image_lc=((count-size_image_lc_gz)/count)*100#lm嵌密壓縮率(%)        
        ws.append(["output{:08d}".format(i),
                   count,
                   size_lc,
                   size_lc_gz,
                   str(round(compress_lc,3))+' %',
                   str(round(compress_code_lc,3))+' %',
                   size_image_lc,
                   size_image_lc_gz,
                   str(round(compress_image_lc,3))+' %',
                   str(round(compress_code_image_lc,3))+' %'])                   
        f_lc.close()             
    wb.save("./embed_mod3/embed_mod3_capacity.xlsx")#寫檔後存檔
tStart = time.time()#計時開始 
embed_ratio=int(input("embedding ratio = "))
cal_capacity("embed_mod3",5000,3,embed_ratio)
tEnd = time.time()#計時結束
wb = openpyxl.load_workbook("./embed_mod3/embed_mod3_capacity.xlsx")
ws = wb['Sheet']
ws.append(["total time",tEnd-tStart])
wb.save("./embed_mod3/embed_mod3_capacity.xlsx")#寫檔後存檔   
print(tEnd-tStart)
