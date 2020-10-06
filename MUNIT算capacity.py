# -*- coding: utf-8 -*-
#######################################################
'''
input
    輸入一張圖片：
        1.資料夾名稱
        2.檔案名稱(圖片)，單純用來記錄在xlsx檔案中
        3.輸出路徑-xlsx
        4.輸出路徑-lc-txt
        5.輸出路徑-lc-png
        6.嵌密mod值
        7.輸出路徑-lc-txt-壓縮
        8.輸出路徑-lc-png-壓縮
        9.嵌密率
output
    產生輸入圖片的xlsx檔(依序將所有圖片的資料寫入xlsx檔中)        
'''
#######################################################
from skimage import io
from openpyxl import Workbook
import openpyxl
import os
import math
def cal_capacity(folder_name,
                 file_name,
                 dir_out,
                 dir_lc,
                 dir_image_lc,
                 num_mod,
                 dir_lc_gz,
                 dir_image_lc_gz):
    f_lc= open(dir_lc,'r')                                      #打開location map.txt來計算capacity
    image_lc = io.imread(dir_image_lc)
    count = 0
    b=int(input("embedding ratio = "))
    for row in range(image_lc.shape[0]):                        #計算capacity，非白區域可嵌密，三個channel，對mod以2為底取log(單位：bit)
        for col in range(image_lc.shape[1]):
            bit=f_lc.read(1)
            if bit== "0" or bit== "2" :
                count+=1
    count*=3*math.log(num_mod,2)*embed_ratio/100                                #capacity
    try:#若檔案存在，直接打開繼續寫xlsx檔
        wb = openpyxl.load_workbook(dir_out)
    except:#若檔案不存在，建立新xlsx檔
        wb = Workbook()
        ws = wb.active
        ws.append([folder_name,"mod="+str(num_mod)])
        ws.append(["","","文字檔","","","","圖片檔"])
        ws.append(["檔名","嵌密量","大小(bit)","壓縮大小","檔案壓縮","嵌密壓縮","大小(bit)","壓縮大小","檔案壓縮","嵌密壓縮"])
        wb.save(dir_out)
        wb = openpyxl.load_workbook(dir_out)
    size_lc=os.path.getsize(dir_lc)*8#lc(txt)大小
    size_image_lc=os.path.getsize(dir_image_lc)*8#lm(png)大小
    size_lc_gz=os.path.getsize(dir_lc_gz)*8#lc壓縮大小
    size_image_lc_gz=os.path.getsize(dir_image_lc_gz)*8#lm壓縮大小
    compress_lc=((size_lc-size_lc_gz)/size_lc)*100
    compress_image_lc=((size_image_lc-size_image_lc_gz)/size_image_lc)*100
    compress_code_lc=((count-size_lc_gz)/count)*100#lc嵌密壓縮率(%)
    compress_code_image_lc=((count-size_image_lc_gz)/count)*100#lm嵌密壓縮率(%)
    ws = wb.get_sheet_by_name('Sheet')
    ws.append([file_name,
               count,
               size_lc,
               size_lc_gz,
               str(round(compress_lc,3))+' %',
               str(round(compress_code_lc,3))+' %',
               size_image_lc,
               size_image_lc_gz,
               str(round(compress_image_lc,3))+' %',
               str(round(compress_code_image_lc,3))+' %'])
    wb.save(dir_out)#寫檔後存檔
    f_lc.close()
'''
cal_capacity("oneshoe1"
             ,"output00000000"
             ,"./oneshoe1/oneshoe1_capacity.xlsx"
             ,"./oneshoe1/output00000000/output00000000_location map.txt"
             ,"./oneshoe1/output00000000/output00000000_lc.png"
             ,5
             ,"./oneshoe1/output00000000/output00000000_location map.tar.gz"
             ,"./oneshoe1/output00000000/output00000000_lc.tar.gz")
'''
for i in range(3):                                                                              #依次執行函式
    for j in range(20):
        name1 = "output{:08d}".format(j)
        name2 = "/output{:08d}".format(j)
        name3 = "oneshoe{}".format(i+1)
        name4 = "./oneshoe{}".format(i+1)
        cal_capacity(name3
             ,name1
             ,name4+'/'+name3+"_capacity.xlsx"
             ,name4+name2+name2+"_location map.txt"
             ,name4+name2+name2+"_lc.png"
             ,7
             ,name4+name2+name2+"_location map.tar.gz"
             ,name4+name2+name2+"_lc.tar.gz"
             )
